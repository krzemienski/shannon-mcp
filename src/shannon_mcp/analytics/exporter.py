"""
Metrics Exporter for Analytics Engine.

Exports analytics data to various formats and destinations.
"""

import json
import csv
import asyncio
import aiofiles
from pathlib import Path
from datetime import datetime, timezone
from typing import List, Dict, Any, Optional, AsyncIterator
from dataclasses import dataclass
from enum import Enum
import zipfile
import tempfile
import shutil

from ..utils.logging import get_logger
from ..utils.errors import ShannonError
from .parser import MetricsParser, ParsedMetric
from .aggregator import MetricsAggregator, AggregationType
from .reporter import ReportGenerator, ReportFormat

logger = get_logger(__name__)


class ExportFormat(str, Enum):
    """Supported export formats."""
    JSON = "json"
    CSV = "csv"
    JSONL = "jsonl"
    PARQUET = "parquet"  # Requires pyarrow
    EXCEL = "excel"  # Requires openpyxl
    ZIP = "zip"  # Archive of multiple formats


@dataclass
class ExportOptions:
    """Options for exporting metrics."""
    format: ExportFormat
    include_raw_metrics: bool = True
    include_aggregations: bool = True
    include_reports: bool = True
    aggregation_types: List[AggregationType] = None
    report_formats: List[ReportFormat] = None
    compress: bool = False
    
    def __post_init__(self):
        """Set defaults."""
        if self.aggregation_types is None:
            self.aggregation_types = [
                AggregationType.DAILY,
                AggregationType.BY_TOOL,
                AggregationType.BY_SESSION
            ]
        
        if self.report_formats is None:
            self.report_formats = [
                ReportFormat.MARKDOWN,
                ReportFormat.JSON
            ]


class MetricsExporter:
    """Exports metrics data to various formats."""
    
    def __init__(
        self,
        parser: MetricsParser,
        aggregator: MetricsAggregator,
        reporter: ReportGenerator
    ):
        """
        Initialize exporter.
        
        Args:
            parser: Metrics parser instance
            aggregator: Metrics aggregator instance
            reporter: Report generator instance
        """
        self.parser = parser
        self.aggregator = aggregator
        self.reporter = reporter
        
    async def export(
        self,
        output_path: Path,
        start_time: datetime,
        end_time: datetime,
        options: ExportOptions,
        filters: Optional[Dict[str, Any]] = None
    ) -> Path:
        """
        Export metrics data.
        
        Args:
            output_path: Output file path
            start_time: Start of time range
            end_time: End of time range
            options: Export options
            filters: Optional filters
            
        Returns:
            Path to exported file
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        if options.format == ExportFormat.JSON:
            return await self._export_json(
                output_path, start_time, end_time, options, filters
            )
        elif options.format == ExportFormat.CSV:
            return await self._export_csv(
                output_path, start_time, end_time, options, filters
            )
        elif options.format == ExportFormat.JSONL:
            return await self._export_jsonl(
                output_path, start_time, end_time, options, filters
            )
        elif options.format == ExportFormat.PARQUET:
            return await self._export_parquet(
                output_path, start_time, end_time, options, filters
            )
        elif options.format == ExportFormat.EXCEL:
            return await self._export_excel(
                output_path, start_time, end_time, options, filters
            )
        elif options.format == ExportFormat.ZIP:
            return await self._export_zip(
                output_path, start_time, end_time, options, filters
            )
        else:
            raise ValueError(f"Unsupported export format: {options.format}")
    
    async def _export_json(
        self,
        output_path: Path,
        start_time: datetime,
        end_time: datetime,
        options: ExportOptions,
        filters: Optional[Dict[str, Any]]
    ) -> Path:
        """Export as JSON."""
        data = {}
        
        # Metadata
        data["metadata"] = {
            "exported_at": datetime.now(timezone.utc).isoformat(),
            "start_time": start_time.isoformat(),
            "end_time": end_time.isoformat(),
            "filters": filters or {}
        }
        
        # Raw metrics
        if options.include_raw_metrics:
            metrics = []
            async for batch in self.parser.stream_metrics(start_time, end_time):
                for metric in batch:
                    if self._apply_filters(metric, filters):
                        metrics.append(metric.entry.to_dict())
            data["metrics"] = metrics
        
        # Aggregations
        if options.include_aggregations:
            data["aggregations"] = {}
            for agg_type in options.aggregation_types:
                result = await self.aggregator.aggregate(
                    agg_type, start_time, end_time, filters
                )
                data["aggregations"][agg_type.value] = result.to_dict()
        
        # Reports
        if options.include_reports:
            data["reports"] = {}
            for report_format in options.report_formats:
                if report_format == ReportFormat.JSON:
                    continue  # Skip JSON in JSON
                    
                report = await self.reporter.generate_report(
                    AggregationType.DAILY,
                    start_time,
                    end_time,
                    report_format,
                    filters
                )
                data["reports"][report_format.value] = report.content
        
        # Write to file
        async with aiofiles.open(output_path, 'w') as f:
            await f.write(json.dumps(data, indent=2))
        
        # Compress if requested
        if options.compress:
            return await self._compress_file(output_path)
        
        return output_path
    
    async def _export_csv(
        self,
        output_path: Path,
        start_time: datetime,
        end_time: datetime,
        options: ExportOptions,
        filters: Optional[Dict[str, Any]]
    ) -> Path:
        """Export as CSV."""
        # For CSV, we'll create multiple files if needed
        base_path = output_path.with_suffix('')
        files_created = []
        
        # Raw metrics CSV
        if options.include_raw_metrics:
            metrics_path = Path(f"{base_path}_metrics.csv")
            
            # Collect all metrics first to determine headers
            all_metrics = []
            headers = set(["id", "timestamp", "type", "session_id", "user_id"])
            
            async for batch in self.parser.stream_metrics(start_time, end_time):
                for metric in batch:
                    if self._apply_filters(metric, filters):
                        all_metrics.append(metric)
                        # Collect all possible data keys
                        headers.update(metric.entry.data.keys())
            
            # Write CSV
            headers = sorted(list(headers))
            
            async with aiofiles.open(metrics_path, 'w') as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                await f.write(','.join(headers) + '\n')
                
                for metric in all_metrics:
                    row = {
                        "id": metric.entry.id,
                        "timestamp": metric.entry.timestamp.isoformat(),
                        "type": metric.entry.type.value,
                        "session_id": metric.entry.session_id or "",
                        "user_id": metric.entry.user_id or ""
                    }
                    # Add data fields
                    for key in headers:
                        if key in metric.entry.data:
                            value = metric.entry.data[key]
                            if isinstance(value, (list, dict)):
                                value = json.dumps(value)
                            row[key] = value
                    
                    # Write row
                    line = ','.join(str(row.get(h, '')) for h in headers)
                    await f.write(line + '\n')
            
            files_created.append(metrics_path)
        
        # Aggregations CSV
        if options.include_aggregations:
            for agg_type in options.aggregation_types:
                agg_path = Path(f"{base_path}_{agg_type.value}.csv")
                
                result = await self.aggregator.aggregate(
                    agg_type, start_time, end_time, filters
                )
                
                # Write time series data as CSV
                if result.time_series:
                    headers = list(result.time_series[0].keys())
                    
                    async with aiofiles.open(agg_path, 'w') as f:
                        await f.write(','.join(headers) + '\n')
                        
                        for entry in result.time_series:
                            values = []
                            for h in headers:
                                value = entry.get(h, '')
                                if isinstance(value, (list, dict)):
                                    value = json.dumps(value)
                                values.append(str(value))
                            await f.write(','.join(values) + '\n')
                    
                    files_created.append(agg_path)
        
        # If multiple files, create a ZIP
        if len(files_created) > 1:
            zip_path = output_path.with_suffix('.zip')
            with zipfile.ZipFile(zip_path, 'w') as zf:
                for file_path in files_created:
                    zf.write(file_path, file_path.name)
                    file_path.unlink()  # Remove individual file
            return zip_path
        
        return files_created[0] if files_created else output_path
    
    async def _export_jsonl(
        self,
        output_path: Path,
        start_time: datetime,
        end_time: datetime,
        options: ExportOptions,
        filters: Optional[Dict[str, Any]]
    ) -> Path:
        """Export as JSONL."""
        async with aiofiles.open(output_path, 'w') as f:
            # Write metadata as first line
            metadata = {
                "_metadata": True,
                "exported_at": datetime.now(timezone.utc).isoformat(),
                "start_time": start_time.isoformat(),
                "end_time": end_time.isoformat(),
                "filters": filters or {}
            }
            await f.write(json.dumps(metadata) + '\n')
            
            # Stream metrics
            if options.include_raw_metrics:
                async for batch in self.parser.stream_metrics(start_time, end_time):
                    for metric in batch:
                        if self._apply_filters(metric, filters):
                            await f.write(
                                json.dumps(metric.entry.to_dict()) + '\n'
                            )
        
        # Compress if requested
        if options.compress:
            return await self._compress_file(output_path)
        
        return output_path
    
    async def _export_parquet(
        self,
        output_path: Path,
        start_time: datetime,
        end_time: datetime,
        options: ExportOptions,
        filters: Optional[Dict[str, Any]]
    ) -> Path:
        """Export as Parquet (requires pyarrow)."""
        try:
            import pyarrow as pa
            import pyarrow.parquet as pq
        except ImportError:
            raise ShannonError(
                "Parquet export requires 'pyarrow' package. "
                "Install with: pip install pyarrow"
            )
        
        # Collect metrics
        records = []
        
        async for batch in self.parser.stream_metrics(start_time, end_time):
            for metric in batch:
                if self._apply_filters(metric, filters):
                    record = {
                        "id": metric.entry.id,
                        "timestamp": metric.entry.timestamp,
                        "type": metric.entry.type.value,
                        "session_id": metric.entry.session_id,
                        "user_id": metric.entry.user_id,
                        "data": json.dumps(metric.entry.data),
                        "metadata": json.dumps(metric.entry.metadata)
                    }
                    
                    # Flatten common fields
                    if metric.tool_name:
                        record["tool_name"] = metric.tool_name
                    if metric.duration_ms is not None:
                        record["duration_ms"] = metric.duration_ms
                    if metric.success is not None:
                        record["success"] = metric.success
                    if metric.token_count is not None:
                        record["token_count"] = metric.token_count
                    
                    records.append(record)
        
        # Create table and write
        if records:
            table = pa.Table.from_pylist(records)
            pq.write_table(table, output_path, compression='snappy')
        else:
            # Write empty table
            schema = pa.schema([
                pa.field("id", pa.string()),
                pa.field("timestamp", pa.timestamp('us', tz='UTC')),
                pa.field("type", pa.string()),
                pa.field("session_id", pa.string()),
                pa.field("user_id", pa.string()),
                pa.field("data", pa.string()),
                pa.field("metadata", pa.string())
            ])
            table = pa.Table.from_pylist([], schema=schema)
            pq.write_table(table, output_path)
        
        return output_path
    
    async def _export_excel(
        self,
        output_path: Path,
        start_time: datetime,
        end_time: datetime,
        options: ExportOptions,
        filters: Optional[Dict[str, Any]]
    ) -> Path:
        """Export as Excel (requires openpyxl)."""
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ShannonError(
                "Excel export requires 'openpyxl' package. "
                "Install with: pip install openpyxl"
            )
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Metadata sheet
        ws_meta = wb.active
        ws_meta.title = "Metadata"
        ws_meta.append(["Field", "Value"])
        ws_meta.append(["Exported At", datetime.now(timezone.utc).isoformat()])
        ws_meta.append(["Start Time", start_time.isoformat()])
        ws_meta.append(["End Time", end_time.isoformat()])
        ws_meta.append(["Filters", json.dumps(filters or {})])
        
        # Raw metrics sheet
        if options.include_raw_metrics:
            ws_metrics = wb.create_sheet("Metrics")
            
            # Headers
            headers = [
                "ID", "Timestamp", "Type", "Session ID", "User ID",
                "Tool Name", "Duration (ms)", "Success", "Tokens", "Data"
            ]
            ws_metrics.append(headers)
            
            # Data
            row_num = 2
            async for batch in self.parser.stream_metrics(start_time, end_time):
                for metric in batch:
                    if self._apply_filters(metric, filters):
                        ws_metrics.append([
                            metric.entry.id,
                            metric.entry.timestamp.isoformat(),
                            metric.entry.type.value,
                            metric.entry.session_id or "",
                            metric.entry.user_id or "",
                            metric.tool_name or "",
                            metric.duration_ms or "",
                            str(metric.success) if metric.success is not None else "",
                            metric.token_count or "",
                            json.dumps(metric.entry.data)
                        ])
                        row_num += 1
            
            # Auto-adjust column widths
            for column_cells in ws_metrics.columns:
                length = max(len(str(cell.value or "")) for cell in column_cells)
                ws_metrics.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)
        
        # Summary sheet
        if options.include_aggregations:
            ws_summary = wb.create_sheet("Summary")
            
            # Get daily aggregation
            result = await self.aggregator.aggregate(
                AggregationType.DAILY, start_time, end_time, filters
            )
            
            # Summary stats
            ws_summary.append(["Metric", "Value"])
            ws_summary.append(["Total Metrics", result.total_metrics])
            ws_summary.append(["Total Sessions", result.total_sessions])
            ws_summary.append(["Total Users", result.total_users])
            ws_summary.append(["Total Errors", result.total_errors])
            ws_summary.append(["Success Rate", f"{result.success_rate:.1%}"])
            ws_summary.append(["Total Tokens", result.total_tokens])
            ws_summary.append([""])
            
            # Tool usage
            ws_summary.append(["Top Tools", ""])
            ws_summary.append(["Tool", "Uses", "Success Rate"])
            for tool, stats in sorted(
                result.tools_usage.items(),
                key=lambda x: x[1]["count"],
                reverse=True
            )[:10]:
                success_rate = stats["success"] / stats["count"] if stats["count"] > 0 else 0
                ws_summary.append([tool, stats["count"], f"{success_rate:.1%}"])
        
        # Save workbook
        wb.save(output_path)
        
        return output_path
    
    async def _export_zip(
        self,
        output_path: Path,
        start_time: datetime,
        end_time: datetime,
        options: ExportOptions,
        filters: Optional[Dict[str, Any]]
    ) -> Path:
        """Export as ZIP containing multiple formats."""
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            files_to_archive = []
            
            # Export different formats
            formats_to_export = [
                (ExportFormat.JSON, "data.json"),
                (ExportFormat.CSV, "data.csv"),
                (ExportFormat.JSONL, "data.jsonl")
            ]
            
            for format_type, filename in formats_to_export:
                try:
                    format_options = ExportOptions(
                        format=format_type,
                        include_raw_metrics=options.include_raw_metrics,
                        include_aggregations=options.include_aggregations,
                        include_reports=options.include_reports,
                        aggregation_types=options.aggregation_types,
                        report_formats=options.report_formats,
                        compress=False  # Don't compress individual files
                    )
                    
                    file_path = await self.export(
                        temp_path / filename,
                        start_time,
                        end_time,
                        format_options,
                        filters
                    )
                    
                    # Handle multiple files (e.g., CSV might create several)
                    if file_path.exists():
                        files_to_archive.append(file_path)
                    
                    # Check for related files
                    base_name = file_path.stem
                    for related in temp_path.glob(f"{base_name}*"):
                        if related not in files_to_archive:
                            files_to_archive.append(related)
                            
                except Exception as e:
                    logger.warning(f"Failed to export {format_type}: {e}")
            
            # Add reports
            if options.include_reports:
                reports_dir = temp_path / "reports"
                reports_dir.mkdir(exist_ok=True)
                
                for report_format in options.report_formats:
                    try:
                        report = await self.reporter.generate_report(
                            AggregationType.DAILY,
                            start_time,
                            end_time,
                            report_format,
                            filters
                        )
                        
                        ext = {
                            ReportFormat.JSON: ".json",
                            ReportFormat.MARKDOWN: ".md",
                            ReportFormat.HTML: ".html",
                            ReportFormat.CSV: ".csv",
                            ReportFormat.TEXT: ".txt"
                        }.get(report_format, ".txt")
                        
                        report_path = reports_dir / f"report_{report_format.value}{ext}"
                        report.save(report_path)
                        files_to_archive.append(report_path)
                        
                    except Exception as e:
                        logger.warning(f"Failed to generate {report_format} report: {e}")
            
            # Create ZIP archive
            with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                for file_path in files_to_archive:
                    arcname = file_path.relative_to(temp_path)
                    zf.write(file_path, arcname)
            
            return output_path
    
    async def _compress_file(self, file_path: Path) -> Path:
        """Compress a file using gzip."""
        import gzip
        
        compressed_path = file_path.with_suffix(file_path.suffix + '.gz')
        
        async with aiofiles.open(file_path, 'rb') as f_in:
            content = await f_in.read()
        
        async with aiofiles.open(compressed_path, 'wb') as f_out:
            compressed = gzip.compress(content, compresslevel=6)
            await f_out.write(compressed)
        
        # Remove original
        file_path.unlink()
        
        return compressed_path
    
    def _apply_filters(
        self,
        metric: ParsedMetric,
        filters: Optional[Dict[str, Any]]
    ) -> bool:
        """Apply filters to a metric."""
        if not filters:
            return True
        
        if "session_id" in filters and metric.session_id != filters["session_id"]:
            return False
        if "user_id" in filters and metric.user_id != filters["user_id"]:
            return False
        if "type" in filters and metric.type != filters["type"]:
            return False
        if "tool_name" in filters and metric.tool_name != filters["tool_name"]:
            return False
        if "agent_id" in filters and metric.agent_id != filters["agent_id"]:
            return False
        
        return True